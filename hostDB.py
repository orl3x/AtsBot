from openpyxl import *
import os
import pyautogui as pag
thisPC = os.environ["COMPUTERNAME"]
#thisPC = "ATS1-4"

wb = load_workbook(filename="ATSbotDB.xlsx", read_only=False)
ws = wb.active

model = "EUG150S350"

def printValues():
    print("Host -> ATS Type")
    for row in range(2, ws.max_row+1):
        currentRow = "A"+str(row)
        atsModel = "B"+str(row)
        print("{} -> {}".format(ws[currentRow].value,int(ws[atsModel].value)))

def findHost(hostName):
    for row in range(2,ws.max_row+1):
        #look for host
        hostRow = "A"+str(row)
        if ws[hostRow].value == hostName:
            return hostRow
    return None

def writeModelInDB(modelName, hostName):
  if findHost(hostName) is not None:
    currentRow ="C"+findHost(hostName)[-1]
    ws[currentRow] = modelName
    wb.save("ATSbotDB.xlsx")
    print("File saved")
    return True
  else:
    return False

def writeModelInDB(modelName):
  if findHost(thisPC) is not None:
    currentRow ="C"+findHost(thisPC)[-1]
    ws[currentRow] = modelName
    wb.save("ATSbotDB.xlsx")
    print("Added "+modelName)
    print("File saved")
    return True
  else:
    print('Error')
    return False

def getModelFromDB(hostName):
    if findHost(hostName) is not None:
        modelRow = "C"+findHost(hostName)[-1]
        return ws[modelRow].value
    else:
        return None

def getModelFromDB():
    if findHost(thisPC) is not None:
        modelRow = "C"+findHost(thisPC)[-1]
        return ws[modelRow].value
    else:
        return None


def getAtsTypeFromDB(hostName):
    if findHost(hostName) is not None:
        modelRow = "B"+findHost(hostName)[-1]
        return ws[modelRow].value
    else:
        return None

def getAtsTypeFromDB():
    if findHost(thisPC) is not None:
        modelRow = "B"+findHost(thisPC)[-1]
        return ws[modelRow].value
    else:
        return None







